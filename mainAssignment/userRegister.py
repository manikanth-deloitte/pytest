import openpyxl


class Register:
    def __init__(self, name, email, phone, age, password):
        self.name = name
        self.email = email
        self.phone = phone
        self.age = age
        self.password = password
        self.lst_details =[self.name, self.email, self.phone, self.age, self.password]
        self.path = "\\Users\\amanikanth\\PycharmProjects\\pytest\\mainAssignment\\resources\\movieDetails.xlsx"
        self.wb = openpyxl.load_workbook(self.path)
        self.sh2 = self.wb['Sheet2']

    def user_register(self):
        r = self.sh2.max_row
        c = self.sh2.max_column
        row = r+1
        for col in range(1,c+1):
            self.sh2.cell(row,col).value = self.lst_details[col-1]
        self.wb.save(self.path)
        print("user successfully registered")

    def checkUser(self,name):
        r = self.sh2.max_row
        col=1
        for row in range(1, r + 1):
            user_name = self.sh2.cell(row, col).value
            if name == user_name:
                print("user registered already!")
                self.wb.save(self.path)
                break
        else:
            self.wb.save(self.path)
            self.user_register()


def user_register():
    name = input("enter name: ")
    email = input("enter email-id: ")
    phone = input("enter phone number: ")
    age = input("enter age: ")
    password = input("enter password: ")
    reg = Register(name, email, phone, age, password)
    reg.checkUser(name)
