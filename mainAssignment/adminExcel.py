import openpyxl


class Excel:
    def __init__(self):
        self.path = "\\Users\\amanikanth\\PycharmProjects\\pytest\\mainAssignment\\resources\\movieDetails.xlsx"
        self.wb = openpyxl.load_workbook(self.path)
        self.sh1 = self.wb['Sheet1']

    def writeExcel(self,lst):
        r = self.sh1.max_row
        c = self.sh1.max_column
        for j in range(1,c+1):
            self.sh1.cell(r+1,j).value = lst[j-1]
        self.wb.save(self.path)

    def editExcel(self):
        r = self.sh1.max_row
        c = self.sh1.max_column
        row = 1
        col = 1
        print("movie lists:")
        for i in range(2, r+1):
            val = self.sh1.cell(i, 1).value
            print(i, f".{val}")
        edit_movie = input("enter the movie name to be edited:")

        # editing movie details
        for i in range(1,r+1):
            val = self.sh1.cell(i,1).value
            if edit_movie == val:
                for j in range(1,c+1):
                    edit_val = self.sh1.cell(row, col).value
                    self.sh1.cell(i, j, value=input(f"enter the {edit_val} to be edited: "))
                    col += 1
                break
        self.wb.save(self.path)

    def delExcel(self):
        r = self.sh1.max_row
        print("movie lists:")
        for i in range(2, r):
            val = self.sh1.cell(i, 1).value
            print(i, f".{val}")
        del_movie = input("enter the movie name to be deleted:")

        # deleting movie details
        for i in range(1, r + 1):
            val = self.sh1.cell(i, 1).value
            if del_movie == val:
                self.sh1.delete_rows(i)
                break
        self.wb.save(self.path)




